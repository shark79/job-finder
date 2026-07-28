import datetime
import os
import re

import openpyxl

OUTREACH_SHEET = "Outreach"
ASSETS_SHEET = "Assets"

OUTREACH_HEADERS = [
    "Company",
    "Code",
    "Job Board Link",
    "Platform",
    "Primary Contact Name",
    "Contact Email",
    "Contact Phone",
    "Resume File Link",
    "Date Drafted",
    "Status",
]

ASSETS_HEADERS = [
    "Company",
    "Code",
    "Job Board Link",
    "Platform",
    "Primary Contact Name",
    "Contact Email",
    "Contact Phone",
    "Resume File Link",
    "Status",
]


def init_tracker(path: str) -> str:
    if os.path.exists(path):
        return path

    wb = openpyxl.Workbook()
    outreach = wb.active
    outreach.title = OUTREACH_SHEET
    outreach.append(OUTREACH_HEADERS)

    assets = wb.create_sheet(ASSETS_SHEET)
    assets.append(ASSETS_HEADERS)

    wb.save(path)
    return path


def _load(path: str):
    return openpyxl.load_workbook(path)


def _company_letter(company: str) -> str:
    match = re.search(r"[A-Za-z]", company)
    return match.group(0).upper() if match else "X"


def get_or_create_company_code(company: str, tracker_path: str) -> str:
    init_tracker(tracker_path)
    wb = _load(tracker_path)
    outreach = wb[OUTREACH_SHEET]

    existing_codes = []
    for row in outreach.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        row_company, row_code = row[0], row[1]
        if row_company.strip().lower() == company.strip().lower():
            return row_code
        if row_code:
            existing_codes.append(row_code)

    letter = _company_letter(company)
    max_n = -1
    for code in existing_codes:
        m = re.match(rf"^{letter}(\d+)$", code)
        if m:
            max_n = max(max_n, int(m.group(1)))

    return f"{letter}{max_n + 1}"


def append_outreach_row(
    tracker_path: str,
    company: str,
    job_board_link: str,
    platform: str,
    contact_name: str,
    contact_email: str,
    contact_phone: str,
    resume_file_link: str,
    status: str = "drafted",
) -> dict:
    init_tracker(tracker_path)
    code = get_or_create_company_code(company, tracker_path)

    wb = _load(tracker_path)
    outreach = wb[OUTREACH_SHEET]
    date_drafted = datetime.date.today().isoformat()
    outreach.append(
        [
            company,
            code,
            job_board_link,
            platform,
            contact_name,
            contact_email,
            contact_phone,
            resume_file_link,
            date_drafted,
            status,
        ]
    )
    wb.save(tracker_path)

    return {
        "code": code,
        "row": {
            "Company": company,
            "Code": code,
            "Job Board Link": job_board_link,
            "Platform": platform,
            "Primary Contact Name": contact_name,
            "Contact Email": contact_email,
            "Contact Phone": contact_phone,
            "Resume File Link": resume_file_link,
            "Date Drafted": date_drafted,
            "Status": status,
        },
    }


def _find_outreach_row(outreach, code: str, job_board_link: str):
    for row in outreach.iter_rows(min_row=2):
        if row[1].value == code and row[2].value == job_board_link:
            return row
    return None


def update_outreach_status(tracker_path: str, code: str, job_board_link: str, status: str) -> bool:
    wb = _load(tracker_path)
    outreach = wb[OUTREACH_SHEET]
    row = _find_outreach_row(outreach, code, job_board_link)
    if row is None:
        return False
    row[9].value = status
    wb.save(tracker_path)
    return True


def mark_turned(tracker_path: str, code: str, job_board_link: str) -> dict:
    wb = _load(tracker_path)
    outreach = wb[OUTREACH_SHEET]
    row = _find_outreach_row(outreach, code, job_board_link)
    if row is None:
        raise ValueError(f"No Outreach row found for code={code!r}, job_board_link={job_board_link!r}")

    row[9].value = "turned"

    identity = [cell.value for cell in row[:8]]
    assets = wb[ASSETS_SHEET]
    new_status = "interested to recruit"
    assets.append(identity + [new_status])

    wb.save(tracker_path)

    return dict(zip(ASSETS_HEADERS, identity + [new_status]))
