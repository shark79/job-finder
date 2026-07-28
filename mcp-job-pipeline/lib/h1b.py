import csv
import json
import os
import re
import tempfile
from urllib.parse import urljoin

import httpx
import openpyxl

DOL_PAGE_URL = "https://www.dol.gov/agencies/eta/foreign-labor/performance"
FILE_PATTERN = re.compile(r'href="([^"]*LCA_Dis[a-z]*closure_Data_FY(\d{4})_Q(\d)\.xlsx)"', re.IGNORECASE)

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "h1b")
INDEX_CSV = os.path.join(DATA_DIR, "h1b_index.csv")
META_JSON = os.path.join(DATA_DIR, "meta.json")

INDEX_FIELDS = [
    "employer_name",
    "job_title",
    "soc_title",
    "worksite_city",
    "worksite_state",
    "visa_class",
    "case_status",
    "full_time_position",
    "begin_date",
]

KEEP_STATUSES = {"CERTIFIED", "CERTIFIED-WITHDRAWN"}


def _find_latest_file_url() -> tuple[str, str]:
    resp = httpx.get(DOL_PAGE_URL, timeout=30, follow_redirects=True)
    resp.raise_for_status()

    candidates = []
    for href, year, quarter in FILE_PATTERN.findall(resp.text):
        url = urljoin(DOL_PAGE_URL, href)
        candidates.append((int(year), int(quarter), url))

    if not candidates:
        raise RuntimeError("No LCA_Disclosure_Data file links found on DOL performance page.")

    candidates.sort(key=lambda c: (c[0], c[1]))
    _, _, latest_url = candidates[-1]
    filename = os.path.basename(latest_url)
    return latest_url, filename


def _load_meta() -> dict:
    if not os.path.exists(META_JSON):
        return {}
    with open(META_JSON) as f:
        return json.load(f)


def _save_meta(meta: dict) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(META_JSON, "w") as f:
        json.dump(meta, f, indent=2)


def refresh_h1b_data(force: bool = False) -> dict:
    """Check DOL's OFLC performance page for the latest LCA disclosure quarterly file.
    Downloads and re-indexes locally only if it's newer than what's cached (or force=True)."""
    latest_url, filename = _find_latest_file_url()
    meta = _load_meta()

    if not force and meta.get("source_file") == filename and os.path.exists(INDEX_CSV):
        return {"status": "up_to_date", "source_file": filename, "row_count": meta.get("row_count")}

    os.makedirs(DATA_DIR, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        with httpx.stream("GET", latest_url, timeout=180, follow_redirects=True) as resp:
            resp.raise_for_status()
            with open(tmp_path, "wb") as f:
                for chunk in resp.iter_bytes(chunk_size=1024 * 1024):
                    f.write(chunk)

        row_count = _reindex(tmp_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    meta = {"source_url": latest_url, "source_file": filename, "row_count": row_count}
    _save_meta(meta)
    return {"status": "refreshed", **meta}


def _reindex(xlsx_path: str) -> int:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col_idx = {name: i for i, name in enumerate(header)}

    required = [
        "EMPLOYER_NAME",
        "JOB_TITLE",
        "SOC_TITLE",
        "WORKSITE_CITY",
        "WORKSITE_STATE",
        "VISA_CLASS",
        "CASE_STATUS",
        "FULL_TIME_POSITION",
        "BEGIN_DATE",
    ]
    missing = [c for c in required if c not in col_idx]
    if missing:
        raise RuntimeError(f"Unexpected LCA file schema, missing columns: {missing}")

    row_count = 0
    with open(INDEX_CSV, "w", newline="", encoding="utf-8") as out:
        writer = csv.writer(out)
        writer.writerow(INDEX_FIELDS)
        for row in rows:
            status = (row[col_idx["CASE_STATUS"]] or "").strip().upper()
            if status not in KEEP_STATUSES:
                continue
            writer.writerow(
                [
                    row[col_idx["EMPLOYER_NAME"]],
                    row[col_idx["JOB_TITLE"]],
                    row[col_idx["SOC_TITLE"]],
                    row[col_idx["WORKSITE_CITY"]],
                    row[col_idx["WORKSITE_STATE"]],
                    row[col_idx["VISA_CLASS"]],
                    status,
                    row[col_idx["FULL_TIME_POSITION"]],
                    row[col_idx["BEGIN_DATE"]],
                ]
            )
            row_count += 1

    wb.close()
    return row_count


def search_h1b_sponsors(
    employer: str | None = None,
    job_title_keyword: str | None = None,
    state: str | None = None,
    limit: int = 50,
) -> list[dict]:
    """Query the locally cached LCA index (built by refresh_h1b_data). Returns companies
    matching the given filters, aggregated with filing counts and sample job titles."""
    if not os.path.exists(INDEX_CSV):
        raise RuntimeError("No local H1B index found. Call refresh_h1b_data() first.")

    employer_q = employer.strip().lower() if employer else None
    job_q = job_title_keyword.strip().lower() if job_title_keyword else None
    state_q = state.strip().upper() if state else None

    aggregated: dict[str, dict] = {}

    with open(INDEX_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = (row["employer_name"] or "").strip()
            if not name:
                continue
            if employer_q and employer_q not in name.lower():
                continue
            if job_q and job_q not in (row["job_title"] or "").lower() and job_q not in (
                row["soc_title"] or ""
            ).lower():
                continue
            if state_q and (row["worksite_state"] or "").strip().upper() != state_q:
                continue

            key = name.upper()
            entry = aggregated.setdefault(
                key, {"employer": name, "count": 0, "job_titles": set(), "worksite_states": set()}
            )
            entry["count"] += 1
            if row["job_title"]:
                entry["job_titles"].add(row["job_title"])
            if row["worksite_state"]:
                entry["worksite_states"].add(row["worksite_state"])

    results = sorted(aggregated.values(), key=lambda e: e["count"], reverse=True)[:limit]
    for r in results:
        r["job_titles"] = sorted(r["job_titles"])[:10]
        r["worksite_states"] = sorted(r["worksite_states"])
    return results
